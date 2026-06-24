from io import BytesIO

from openpyxl import Workbook
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.files.base import ContentFile
from django.http import FileResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.db import models, transaction

from .forms import ImportForm, PriceUpdateForm, ProductForm, LeadForm
from .models import ExportFile, ImportFile, MasterProduct, Product, ProductChange, Quotation, Pricing, Notification, Lead
import csv
from datetime import datetime
from decimal import Decimal
from django.utils import timezone
from django.urls import reverse
from django.db.models import Q


User = get_user_model()
from .utils import ADMIN_USERNAME, is_admin, is_sales, is_production

def record_changes(product, old_product, user):
    fields = ["sku", "name", "quantity", "production_price", "profit_percent"]
    for field in fields:
        old_value = getattr(old_product, field)
        new_value = getattr(product, field)
        if old_value != new_value:
            ProductChange.objects.create(
                product=product,
                changed_by=user,
                field_name=field,
                old_value=str(old_value),
                new_value=str(new_value),
            )

@login_required
def get_notifications(request):
    if is_admin(request.user):
        base_qs = Notification.objects.filter(recipient=request.user)
    else:
        base_qs = Notification.objects.filter(models.Q(recipient=request.user) | models.Q(recipient__isnull=True)).distinct()
        
    unread_count = base_qs.exclude(read_by=request.user).count()
    qs = base_qs.prefetch_related('read_by').order_by('-created_at')[:10]

    data = []
    for n in qs:
        is_read_for_user = any(reader.id == request.user.id for reader in n.read_by.all())
        data.append({
            'id': n.id,
            'sender': n.sender.username,
            'message': n.message,
            'is_read': is_read_for_user,
            'created_at': n.created_at.strftime("%b %d, %H:%M")
        })
    return JsonResponse({'notifications': data, 'unread_count': unread_count})

@login_required
def mark_notifications_read(request):
    if request.method == "POST":
        if is_admin(request.user):
            notifications_to_mark = Notification.objects.filter(recipient=request.user).exclude(read_by=request.user)
        else:
            notifications_to_mark = Notification.objects.filter(models.Q(recipient=request.user) | models.Q(recipient__isnull=True)).distinct().exclude(read_by=request.user)
        
        user = request.user
        Notification.read_by.through.objects.bulk_create([
            Notification.read_by.through(notification_id=n_id, user_id=user.id)
            for n_id in notifications_to_mark.values_list('id', flat=True)
        ], ignore_conflicts=True)

        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'})

@login_required
def send_inquiry(request):
    if request.method == "POST":
        message = request.POST.get('message')
        if message:
            if is_admin(request.user):
                Notification.objects.create(sender=request.user, message=message, recipient=None)
            else:
                admin_user = User.objects.filter(username__iexact=ADMIN_USERNAME).first()
                if admin_user:
                    Notification.objects.create(sender=request.user, message=message, recipient=admin_user)
    return redirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def dashboard(request):
    if is_admin(request.user) or request.user.groups.filter(name='Sales').exists():
        return sales_dashboard(request)
    else:
        return production_dashboard(request)

@login_required
def sales_dashboard(request):
    query = (request.GET.get('q') or '').strip()

    if query:
        matching_masters = MasterProduct.objects.filter(
            Q(code__icontains=query) |
            Q(description__icontains=query) |
            Q(specification__icontains=query)
        )
        for mp in matching_masters:
            Product.objects.get_or_create(
                sku=mp.code,
                master=mp,
                defaults={
                    'name': mp.description,
                    'production_price': Decimal("0.00"),
                    'profit_percent': Decimal("0.00"),
                    'quantity': Decimal("0.00"),
                }
            )

    search_results = []
    if query:
        search_results = Product.objects.filter(
            Q(sku__icontains=query) |
            Q(master__code__icontains=query) |
            Q(name__icontains=query) |
            Q(master__description__icontains=query) |
            Q(master__specification__icontains=query)
        ).select_related('master')

    recent_quotations = Quotation.objects.filter(salesperson=request.user).order_by('-created_at')[:10]
    recent_updates = ProductChange.objects.select_related("product").filter(field_name__in=['production_price', 'profit_percent', 'selling_price']).order_by('-created_at')[:5]

    master_products = MasterProduct.objects.all().order_by('code')
    products = Product.objects.all().order_by('sku')

    return render(request, "sales_dashboard.html", {
        "query": query,
        "search_results": search_results,
        "recent_quotations": recent_quotations,
        "recent_updates": recent_updates,
        "master_products": master_products,
        "products": products,
    })

@login_required
def crm_dashboard(request):
    if not is_sales(request.user):
        return redirect('dashboard')
    leads = Lead.objects.filter(salesperson=request.user).order_by('-created_at')
    leads_by_stage = {
        'New': leads.filter(stage='New'),
        'Qualified': leads.filter(stage='Qualified'),
        'Proposition': leads.filter(stage='Proposition'),
        'Won': leads.filter(stage='Won'),
    }
    return render(request, "crm_dashboard.html", {
        "leads_by_stage": leads_by_stage,
    })

@login_required
def create_quotation(request):
    if not is_sales(request.user):
        return redirect('dashboard')
    if request.method == "POST":
        product_id = request.POST.get("product_id")
        customer_name = request.POST.get("customer_name")
        quantity = request.POST.get("quantity")
        discount_percent = request.POST.get("discount_percent") or 0
        
        if product_id and customer_name and quantity:
            product = get_object_or_404(Product, pk=product_id)
            Quotation.objects.create(
                product=product,
                salesperson=request.user,
                customer_name=customer_name,
                quantity=Decimal(str(quantity)),
                discount_percent=Decimal(str(discount_percent))
            )
    return redirect("dashboard")

@login_required
def production_dashboard(request):
    query = (request.GET.get('q') or '').strip()
    
    if query:
        mp = MasterProduct.objects.filter(code=query).first()
        if mp:
            Product.objects.get_or_create(
                sku=mp.code,
                master=mp,
                defaults={
                    'name': mp.description,
                    'production_price': Decimal("0.00"),
                    'profit_percent': Decimal("0.00"),
                    'quantity': Decimal("0.00"),
                }
            )

    search_results = []
    if query:
        search_results = Product.objects.filter(
            Q(sku__icontains=query) |
            Q(master__code__icontains=query) |
            Q(name__icontains=query)
        ).select_related('master')

    # Quick cost update form handling in dashboard
    if request.method == "POST" and "update_cost" in request.POST:
        product_id = request.POST.get("product_id")
        new_cost = request.POST.get("production_price")
        new_profit = request.POST.get("profit_percent")
        if product_id and new_cost is not None and new_profit is not None:
            try:
                product = Product.objects.get(id=product_id)
                old_product = Product.objects.get(id=product_id)
                product.production_price = Decimal(str(new_cost))
                product.profit_percent = Decimal(str(new_profit))
                product.updated_by = request.user
                product.save()
                record_changes(product, old_product, request.user)
                return redirect(request.path + ("?q=" + query if query else ""))
            except Product.DoesNotExist:
                pass

    # Check for prices older than 4 days
    four_days_ago = timezone.now() - timezone.timedelta(days=4)
    old_products = Product.objects.filter(updated_at__lt=four_days_ago)
    admin_user = User.objects.filter(username__iexact=ADMIN_USERNAME).first()
    
    if admin_user and old_products.exists():
        last_24h = timezone.now() - timezone.timedelta(days=1)
        for product in old_products:
            msg = f"Price for {product.sku} ({product.name}) is more than 4 days old. Please update it."
            if not Notification.objects.filter(message=msg, created_at__gte=last_24h).exists():
                Notification.objects.create(
                    sender=admin_user,
                    recipient=None,
                    message=msg
                )

    attention_products = Product.objects.filter(Q(production_price=0) | Q(selling_price=0)).order_by('-updated_at')[:10]
    history = ProductChange.objects.select_related("product", "changed_by").order_by("-created_at")[:10]
    master_products = MasterProduct.objects.all().order_by('code')

    return render(
        request,
        "dashboard.html",
        {
            "query": query,
            "search_results": search_results,
            "attention_products": attention_products,
            "history": history,
            "can_invite": is_admin(request.user),
            "admin_username": ADMIN_USERNAME,
            "master_products": master_products,
        },
    )


@login_required
def history_view(request):
    if not is_production(request.user):
        return redirect('dashboard')
    query = (request.GET.get("q") or "").strip()
    date_value = (request.GET.get("date") or "").strip()

    changes = ProductChange.objects.select_related("product", "changed_by").order_by("-created_at")

    if query:
        changes = changes.filter(
            (
                models.Q(product__sku__icontains=query)
                | models.Q(product__name__icontains=query)
                | models.Q(changed_by__username__icontains=query)
                | models.Q(field_name__icontains=query)
                | models.Q(old_value__icontains=query)
                | models.Q(new_value__icontains=query)
            )
        )

    if date_value:
        parsed_date = parse_date(date_value)
        if parsed_date:
            changes = changes.filter(created_at__date=parsed_date)

    return render(
        request,
        "history.html",
        {
            "changes": changes,
            "query": query,
            "date_value": date_value,
            "total_count": changes.count(),
        },
    )


@login_required
@user_passes_test(is_admin)
def invite_people(request):
    members = User.objects.order_by("username")
    errors = []
    created_user = None
    invite_text = None

    if request.method == "POST":
        action = request.POST.get("action", "create")

        if action == "create":
            username = (request.POST.get("username") or "").strip()
            password = (request.POST.get("password") or "").strip()
            role = request.POST.get("role")

            if not username:
                errors.append("Username is required")
            if not password:
                errors.append("Password is required")
            if role not in ["sales", "production"]:
                errors.append("Role is required and must be sales or production")

            if not errors:
                if User.objects.filter(username__iexact=username).exists():
                    errors.append("That username already exists")
                else:
                    created_user = User.objects.create_user(username=username, password=password)
                    if role == "sales":
                        group, _ = Group.objects.get_or_create(name="Sales")
                        created_user.groups.add(group)
                    elif role == "production":
                        group, _ = Group.objects.get_or_create(name="Production")
                        created_user.groups.add(group)

                    invite_text = (
                        "Join our company's private collaborative Excel tool.\n\n"
                        f"Login URL: {request.build_absolute_uri('/accounts/login/')}\n"
                        f"Username: {created_user.username}\n"
                        f"Password: {password}\n\n"
                        "Use this account to access the workspace."
                    )

        elif action == "delete":
            username = (request.POST.get("username") or "").strip()
            if not username:
                errors.append("Username is required")
            elif username.lower() == ADMIN_USERNAME.lower():
                errors.append("The admin account cannot be removed")
            else:
                User.objects.filter(username__iexact=username).delete()

        members = User.objects.order_by("username")

    return render(
        request,
        "invite_people.html",
        {
            "members": members,
            "errors": errors,
            "created_user": created_user,
            "invite_text": invite_text,
            "admin_username": ADMIN_USERNAME,
        },
    )


@login_required
def import_detail(request, import_id):
    if not is_production(request.user):
        return redirect('dashboard')
    imp = get_object_or_404(ImportFile, pk=import_id)
    fpath = imp.file.path
    with open(fpath, "rb") as fh:
        rows = list(iter_rows_from_file(fh, imp.file.name))
    parsed, errors = validate_and_parse_rows(rows)
    return render(request, "import_preview.html", {"import": imp, "rows": parsed, "errors": errors})


@login_required
def products_list(request):
    errors = []
    edit_id = None
    can_edit = is_admin(request.user) or request.user.groups.filter(name='Production').exists()
    query = (request.GET.get('q') or '').strip()

    if request.method == 'POST' and can_edit:
        action = request.POST.get('action')
        description = (request.POST.get('description') or '').strip()
        specifications = (request.POST.get('specifications') or '').strip()
        approval = (request.POST.get('approval') or '').strip()
        product_id = request.POST.get('product_id')

        if action == 'add':
            viscosity = (request.POST.get('viscosity') or '').strip()
            liters_per_case_str = (request.POST.get('liters_per_case') or '').strip()
            liters_per_case = None
            if liters_per_case_str:
                try:
                    liters_per_case = Decimal(liters_per_case_str)
                except Exception:
                    errors.append('Litres per case must be a valid number.')

            if not description:
                errors.append('Description is required.')
            
            codes_and_packs = [
                (request.POST.get('code_12x1', '').strip(), '12x1'),
                (request.POST.get('code_4x4', '').strip(), '4x4'),
                (request.POST.get('code_4x5', '').strip(), '4x5'),
                (request.POST.get('code_20', '').strip(), '20'),
                (request.POST.get('code_208', '').strip(), '208'),
            ]
            
            valid_codes = []
            for code, pack in codes_and_packs:
                if code:
                    if not code.isdigit():
                        errors.append(f'Product code {code} for {pack} must contain only digits.')
                    elif MasterProduct.objects.filter(code=code, description=description, packaging=pack).exists():
                        errors.append(f'Product with code {code}, name {description} and packaging {pack} already exists.')
                    else:
                        valid_codes.append((code, pack))
            
            if not valid_codes and not errors:
                errors.append('At least one product code must be provided.')

            if not errors:
                for code, pack in valid_codes:
                    MasterProduct.objects.create(
                        code=code, description=description, viscosity=viscosity, specification=specifications,
                        approval=approval, packaging=pack, liters_per_case=liters_per_case
                    )
                return redirect('products_list')

        if action == 'edit':
            edit_id = request.POST.get('product_id')

        if action == 'update':
            code = (request.POST.get('code') or '').strip()
            packaging = (request.POST.get('packaging') or '').strip()
            viscosity = (request.POST.get('viscosity') or '').strip()
            liters_per_case_str = (request.POST.get('liters_per_case') or '').strip()
            liters_per_case = None
            if liters_per_case_str:
                try:
                    liters_per_case = Decimal(liters_per_case_str)
                except Exception:
                    errors.append('Litres per case must be a valid number.')

            if not code or not code.isdigit():
                errors.append('Product code must contain only digits.')
            if not description:
                errors.append('Description is required.')
                
            if not errors:
                try:
                    master = MasterProduct.objects.get(id=product_id)
                    if (master.code != code or master.description != description or master.packaging != packaging) and MasterProduct.objects.filter(code=code, description=description, packaging=packaging).exists():
                        errors.append(f'Product with code {code}, name {description} and packaging {packaging} already exists.')
                    else:
                        master.code = code
                        master.description = description
                        master.viscosity = viscosity
                        master.specification = specifications
                        master.approval = approval
                        master.packaging = packaging
                        master.liters_per_case = liters_per_case
                        master.save()
                        return redirect('products_list')
                except MasterProduct.DoesNotExist:
                    errors.append('Product not found.')

        if action == 'delete':
            try:
                MasterProduct.objects.get(id=product_id).delete()
                return redirect('products_list')
            except MasterProduct.DoesNotExist:
                errors.append('Product not found.')

    rows = MasterProduct.objects.all().order_by('code')
    if query:
        rows = rows.filter(
            models.Q(description__icontains=query) |
            models.Q(code__icontains=query) |
            models.Q(specification__icontains=query)
        )
    form_values = {}
    if edit_id:
        try:
            master = MasterProduct.objects.get(id=edit_id)
            form_values = {
                'product_id': master.id,
                'code': master.code,
                'description': master.description,
                'specifications': master.specification,
                'approval': master.approval,
                'packaging': master.packaging,
            }
        except MasterProduct.DoesNotExist:
            pass

    return render(request, 'products.html', {'rows': rows, 'errors': errors, 'form': form_values, 'can_edit': can_edit, 'query': query})

@login_required
def pricing_list(request):
    if not is_production(request.user):
        return redirect('dashboard')
    errors = []
    edit_id = None
    can_edit = is_admin(request.user) or request.user.groups.filter(name='Production').exists()

    if request.method == 'POST' and can_edit:
        action = request.POST.get('action')
        product_id = request.POST.get('product')
        packaging = request.POST.get('packaging') or ''
        production_cost_per_liter = request.POST.get('production_cost_per_liter') or 0
        quantity = request.POST.get('quantity') or 0
        discount = request.POST.get('discount') or 0
        pricing_id = request.POST.get('pricing_id')

        final_cost = Decimal("0")
        selling_price = 0
        price = 0

        if action in ('add', 'update'):
            if not product_id:
                errors.append('Product is required.')
            else:
                try:
                    master = MasterProduct.objects.get(id=product_id)
                    import re
                    packaging_master = (master.packaging or "").strip().lower()
                    packaging_master = re.sub(r'l$', '', packaging_master).strip()
                    
                    pack_multiplier = Decimal("0")
                    if packaging_master == '12x1':
                        pack_multiplier = Decimal("12")
                    elif packaging_master == '4x4':
                        pack_multiplier = Decimal("16")
                    elif packaging_master == '4x5':
                        pack_multiplier = Decimal("20")
                    else:
                        match = re.match(r"^(\d+)\s*x\s*(\d+)$", packaging_master)
                        if match:
                            pack_multiplier = Decimal(match.group(1)) * Decimal(match.group(2))
                        else:
                            match_single = re.search(r"^\d+", packaging_master)
                            if match_single:
                                pack_multiplier = Decimal(match_single.group())

                    try:
                        match = re.search(r"[-+]?\d*\.\d+|\d+", packaging)
                        packaging_val = Decimal(match.group()) if match else Decimal("0")
                    except Exception:
                        packaging_val = Decimal("0")
                    
                    prod_cost = Decimal(str(production_cost_per_liter)) if production_cost_per_liter else Decimal("0")
                    final_cost = (prod_cost * pack_multiplier + packaging_val).quantize(Decimal("0.01"))
                    selling_price = (final_cost + (final_cost * Decimal("0.20"))).quantize(Decimal("0.01"))
                    
                    qty_val = Decimal(str(quantity)) if quantity else Decimal("0")
                    discount_val = Decimal(str(discount)) if discount else Decimal("0")
                    price = ((selling_price * qty_val) - discount_val).quantize(Decimal("0.01"))
                except MasterProduct.DoesNotExist:
                    pass

        if action == 'add' and not errors:
            try:
                master = MasterProduct.objects.get(id=product_id)
                Pricing.objects.create(
                    product=master,
                    packaging=packaging,
                    production_cost_per_liter=production_cost_per_liter,
                    quantity=quantity,
                    discount=discount,
                    final_cost=final_cost,
                    selling_price=selling_price,
                    price=price,
                )
                return redirect('pricing_list')
            except MasterProduct.DoesNotExist:
                errors.append("Invalid product selected.")

        if action == 'edit':
            edit_id = request.POST.get('pricing_id')

        if action == 'update' and not errors:
            try:
                pricing = Pricing.objects.get(id=pricing_id)
                master = MasterProduct.objects.get(id=product_id)
                pricing.product = master
                pricing.packaging = packaging
                pricing.production_cost_per_liter = production_cost_per_liter
                pricing.quantity = quantity
                pricing.discount = discount
                pricing.final_cost = final_cost
                pricing.selling_price = selling_price
                pricing.price = price
                pricing.save()
                return redirect('pricing_list')
            except Pricing.DoesNotExist:
                errors.append('Pricing not found.')
            except MasterProduct.DoesNotExist:
                errors.append('Product not found.')

        if action == 'delete':
            try:
                Pricing.objects.get(id=pricing_id).delete()
                return redirect('pricing_list')
            except Pricing.DoesNotExist:
                errors.append('Pricing not found.')

    rows = Pricing.objects.select_related('product').all().order_by('product__code')
    master_products = MasterProduct.objects.all().order_by('code')
    
    form_values = {}
    if edit_id:
        try:
            pricing = Pricing.objects.get(id=edit_id)
            form_values = {
                'pricing_id': pricing.id,
                'product_id': pricing.product.id,
                'production_cost_per_liter': pricing.production_cost_per_liter,
                'quantity': pricing.quantity,
                'discount': pricing.discount,
            }
        except Pricing.DoesNotExist:
            pass

    return render(request, 'pricing.html', {
        'rows': rows, 
        'errors': errors, 
        'form': form_values,
        'master_products': master_products,
        'can_edit': can_edit
    })

@login_required
@user_passes_test(is_admin)
def product_create(request):
    form = ProductForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        product = form.save(commit=False)
        product.updated_by = request.user
        product.save()
        ProductChange.objects.create(product=product, changed_by=request.user, field_name="created", old_value="", new_value=product.sku)
        return redirect("dashboard")
    return render(request, "product_form.html", {"form": form})


@login_required
@user_passes_test(is_admin)
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    old_product = Product.objects.get(pk=pk)
    form = ProductForm(request.POST or None, instance=product)
    if request.method == "POST" and form.is_valid():
        product = form.save(commit=False)
        product.updated_by = request.user
        record_changes(product, old_product, request.user)
        product.save()
        return redirect("dashboard")
    return render(request, "product_form.html", {"form": form})


@login_required
def price_update(request, pk):
    if not is_production(request.user):
        return redirect('dashboard')
    product = get_object_or_404(Product, pk=pk)
    old_product = Product.objects.get(pk=pk)
    if request.user.is_staff:
        form = ProductForm(request.POST or None, instance=product)
    else:
        form = PriceUpdateForm(request.POST or None, instance=product)
    if request.method == "POST" and form.is_valid():
        product = form.save(commit=False)
        product.updated_by = request.user
        record_changes(product, old_product, request.user)
        product.save()
        return redirect("dashboard")
    return render(request, "product_form.html", {"form": form})


@login_required
def export_dashboard(request):
    if not is_production(request.user):
        return redirect('dashboard')
    wb = Workbook()
    ws = wb.active
    ws.title = "Pricing"
    ws.append(["SKU", "Name", "Quantity", "Production Price", "Profit Percent", "Selling Price"])
    for product in Product.objects.order_by("sku"):
        ws.append([
            product.sku,
            product.name,
            float(product.quantity),
            float(product.production_price),
            float(product.profit_percent),
            float(product.selling_price),
        ])
    buffer = BytesIO()
    wb.save(buffer)
    export = ExportFile(created_by=request.user)
    export.file.save("pricing-export.xlsx", ContentFile(buffer.getvalue()), save=True)
    return FileResponse(export.file.open("rb"), as_attachment=True, filename="pricing-export.xlsx")


def parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        pass
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def iter_rows_from_file(fobj, filename):
    name = filename.lower()
    # CSV files (including files uploaded with .csv extension)
    if name.endswith(".csv"):
        text = fobj.read().decode("utf-8-sig")
        reader = csv.DictReader(text.splitlines())
        for row in reader:
            yield row
        return

    # Try openpyxl for modern Excel (.xlsx)
    try:
        from openpyxl import load_workbook

        fobj.seek(0)
        wb = load_workbook(fobj, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.rows)]
        for r in ws.iter_rows(min_row=2, values_only=True):
            yield {headers[i]: r[i] for i in range(len(headers))}
        return
    except Exception:
        pass

    # Fallback: try pandas if available (supports many spreadsheet types)
    try:
        import pandas as pd

        fobj.seek(0)
        # pandas can read from a file-like object for many formats
        df = pd.read_excel(fobj)
        for _, row in df.fillna("").iterrows():
            yield {str(k): (v if not pd.isna(v) else "") for k, v in row.items()}
        return
    except Exception:
        pass

    # Final fallback: try to decode as text and parse as CSV-like
    try:
        fobj.seek(0)
        text = fobj.read().decode("utf-8-sig")
        reader = csv.DictReader(text.splitlines())
        for row in reader:
            yield row
        return
    except Exception:
        return


def validate_and_parse_rows(rows):
    parsed = []
    errors = []
    for idx, row in enumerate(rows, start=1):
        sku = (row.get("sku") or row.get("SKU") or row.get("Sku") or "").strip()
        name = (row.get("name") or row.get("Name") or "").strip()
        qty = row.get("quantity") or row.get("Quantity") or row.get("QTY") or 0
        prod = row.get("production_price") or row.get("Production Price") or row.get("productionprice")
        profit = row.get("profit_percent") or row.get("Profit Percent") or row.get("profitpercent")
        datev = row.get("date") or row.get("effective_date") or row.get("Date")
        row_errors = []
        if not sku:
            row_errors.append("Missing SKU")
        if not name:
            row_errors.append("Missing name")
        try:
            quantity = float(qty) if qty not in (None, "") else 0
        except Exception:
            row_errors.append("Invalid quantity")
            quantity = 0
        try:
            production_price = float(prod) if prod not in (None, "") else 0
        except Exception:
            row_errors.append("Invalid production_price")
            production_price = 0
        try:
            profit_percent = float(profit) if profit not in (None, "") else 0
        except Exception:
            row_errors.append("Invalid profit_percent")
            profit_percent = 0
        eff_date = parse_date(datev)
        # production_price and profit_percent are important
        if prod in (None, ""):
            row_errors.append("Missing production_price")
        if profit in (None, ""):
            row_errors.append("Missing profit_percent")

        if row_errors:
            errors.append({"row": idx, "sku": sku, "errors": row_errors})
        parsed.append({
            "sku": sku,
            "name": name,
            "quantity": quantity,
            "production_price": production_price,
            "profit_percent": profit_percent,
            "selling_price": (Decimal(str(production_price)) + (Decimal(str(production_price)) * Decimal(str(profit_percent)) / Decimal("100"))).quantize(Decimal("0.01")) if production_price is not None else Decimal("0.00"),
            "effective_date": eff_date,
        })
    return parsed, errors


@login_required
def import_upload(request):
    if not is_production(request.user):
        return redirect('dashboard')
    if request.method == "POST":
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES["file"]
            rows = list(iter_rows_from_file(uploaded_file, uploaded_file.name))
            parsed, errors = validate_and_parse_rows(rows)
            
            session_rows = request.session.get("create_excel_rows", [])
            for row in parsed:
                if row.get("sku"):
                    session_rows.append({
                        "sku": row.get("sku", ""),
                        "name": row.get("name", ""),
                        "quantity": float(row.get("quantity") or 0.0),
                        "production_price": float(row.get("production_price") or 0.0),
                        "profit_percent": float(row.get("profit_percent") or 0.0),
                    })
            request.session["create_excel_rows"] = session_rows
            request.session.modified = True
            
            return redirect("products_list")
    else:
        form = ImportForm()
    return render(request, "import_form.html", {"form": form})


@login_required
def import_apply(request, import_id):
    if not is_production(request.user):
        return redirect('dashboard')
    # Allow anonymous application; updated_by set to None if anonymous
    imp = get_object_or_404(ImportFile, pk=import_id)
    fpath = imp.file.path
    with open(fpath, "rb") as fh:
        rows = list(iter_rows_from_file(fh, imp.file.name))
    parsed, errors = validate_and_parse_rows(rows)
    applied = 0
    actor = request.user if getattr(request, "user", None) and request.user.is_authenticated else None
    with transaction.atomic():
        for row in parsed:
            if not row["sku"]:
                continue
            try:
                prod = Product.objects.filter(sku=row["sku"], name=row["name"]).first()
                if prod:
                    prod.name = row["name"] or prod.name
                    prod.quantity = row["quantity"]
                    prod.production_price = row["production_price"]
                    prod.profit_percent = row["profit_percent"]
                    prod.effective_date = row["effective_date"]
                    prod.updated_by = actor
                    prod.save()
                else:
                    master = MasterProduct.objects.filter(code=row["sku"], description=row["name"]).first()
                    Product.objects.create(
                        sku=row["sku"],
                        name=row["name"] or row["sku"],
                        master=master,
                        quantity=row["quantity"],
                        production_price=row["production_price"],
                        profit_percent=row["profit_percent"],
                        effective_date=row["effective_date"],
                        updated_by=actor,
                    )
                applied += 1
            except Exception as e:
                errors.append({"row": row.get("sku"), "errors": [str(e)]})
    imp.processed = True
    imp.error_summary = {"errors": errors, "applied": applied}
    imp.save()
    return redirect(reverse("dashboard"))


@login_required
def download_template(request):
    if not is_production(request.user):
        return redirect('dashboard')
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(["sku", "name", "quantity", "production_price", "profit_percent", "date"])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="import-template.xlsx")

@login_required
def crm_lead_create(request):
    if not is_sales(request.user):
        return redirect('dashboard')
    if request.method == 'POST':
        form = LeadForm(request.POST)
        if form.is_valid():
            lead = form.save(commit=False)
            lead.salesperson = request.user
            lead.save()
    return redirect('crm_dashboard')

@login_required
def crm_lead_update(request, pk):
    if not is_sales(request.user):
        return redirect('dashboard')
    lead = get_object_or_404(Lead, pk=pk)
    if request.method == 'POST':
        new_stage = request.POST.get('stage')
        if new_stage in dict(Lead.STAGE_CHOICES):
            lead.stage = new_stage
            lead.save()
    return redirect('crm_dashboard')


@login_required
def crm_lead_edit(request, pk):
    if not is_sales(request.user):
        return redirect('dashboard')
    lead = get_object_or_404(Lead, pk=pk)
    if request.method == 'POST':
        form = LeadForm(request.POST, instance=lead)
        if form.is_valid():
            form.save()
    return redirect('crm_dashboard')

@login_required
def crm_lead_delete(request, pk):
    if not is_sales(request.user):
        return redirect('dashboard')
    lead = get_object_or_404(Lead, pk=pk)
    if request.method == 'POST':
        lead.delete()
    return redirect('crm_dashboard')

@login_required
def generate_quotation_pdf(request, id):
    if not is_sales(request.user):
        return redirect('dashboard')
    quote = get_object_or_404(Quotation, id=id, salesperson=request.user)
    valid_until = quote.created_at + timezone.timedelta(days=30)
    context = {
        'quote': quote,
        'date': quote.created_at.strftime('%B %d, %Y'),
        'valid_until': valid_until.strftime('%B %d, %Y'),
    }
    return render(request, 'quotation_document.html', context)

@login_required
def generate_proforma_pdf(request, id):
    if not is_sales(request.user):
        return redirect('dashboard')
    quote = get_object_or_404(Quotation, id=id, salesperson=request.user)
    context = {
        'quote': quote,
        'date': quote.created_at.strftime('%B %d, %Y'),
    }
    return render(request, 'proforma_invoice_document.html', context)

@login_required
def generate_lead_quotation(request, pk):
    if not is_sales(request.user):
        return redirect('dashboard')
    lead = get_object_or_404(Lead, pk=pk, salesperson=request.user)
    valid_until = timezone.now() + timezone.timedelta(days=30)
    context = {
        'lead': lead,
        'date': timezone.now().strftime('%B %d, %Y'),
        'valid_until': valid_until.strftime('%B %d, %Y'),
    }
    return render(request, 'quotation_document.html', context)

@login_required
def generate_lead_proforma(request, pk):
    if not is_sales(request.user):
        return redirect('dashboard')
    lead = get_object_or_404(Lead, pk=pk, salesperson=request.user)
    context = {
        'lead': lead,
        'date': timezone.now().strftime('%B %d, %Y'),
    }
    return render(request, 'proforma_invoice_document.html', context)
